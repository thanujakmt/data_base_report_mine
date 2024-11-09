
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from utility.niche_details import sender_email,receiver_email,sender_password,cc_emails
from datetime import date
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
from email.mime.base import MIMEBase
from email import encoders

subject = f"All Niche Database Report On new | {date.today()} |"
subject_usa_per = f"All Niche USA Percintile Database Report On new | {date.today()} |"
subject_10k = f"Niche Business Report for Every 10,000 People in Data Scraping Database On new | {date.today()} |"

def create_excel_with_styles(df):
    # Create a workbook and access the active worksheet
    wb = Workbook()
    ws = wb.active
    
    # Add DataFrame rows to worksheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Style the header row
    grayFill = PatternFill(start_color='00CCCCCC',
                           end_color='00CCCCCC',
                           fill_type='solid')
    
    for cell in ws[1]:  # Assuming the first row is the header
        cell.fill = grayFill

    # Save the workbook to a bytes buffer
    from io import BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file

def email_sender(html_content,subject_as_per_mail,is_business_10k = False):
    
    subject = f"{subject_as_per_mail}"

    # Create the email message
    message = MIMEMultipart()
    message["From"] = sender_email
    message["To"] = receiver_email
    message["Subject"] = subject
    message["Cc"] = ", ".join(cc_emails)

    # Attach HTML content
    html_part = MIMEText(html_content,"html")
    message.attach(html_part)
    
    if is_business_10k:
        df = pd.read_html(html_content)[0]
        excel_buffer = create_excel_with_styles(df)

        # Create a MIME part for the Excel file
        part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        part.set_payload(excel_buffer.getvalue())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', 'attachment; filename="Report.xlsx"')
        message.attach(part)

    # Set up the SMTP server
    smtp_server = "smtp.gmail.com"
    smtp_port = 587

    # Send the email
    with smtplib.SMTP(smtp_server, smtp_port) as server:
        server.starttls()
        server.login(sender_email, sender_password)
        server.send_message(message)
        print("Email sent successfully!")
        